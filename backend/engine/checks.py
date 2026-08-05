"""步驟結果的確定性檢查。

Atlas 這一塊分兩層：LLM 驗證（validator.py，把 output.expect 的自然語言送給
模型判斷）+ 確定性檢查。Atlas-Lite 只有確定性檢查 —— 不猜、不花錢、
每次結果一樣，錯了也講得出是哪個檔哪一點不對。

檢查順序（任一不過就 failed）：
  1. exit code
  2. 宣告的 output.path 存在、非 0 bytes
  3. Office 假檔（副檔名是 .docx 但內容不是 ZIP）
  4. CSV 至少有 header + 1 列 / Excel 至少有一個工作表
真正的結構驗證請用 output.json_schema（見 schema_gate.py）。
"""
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

# OOXML 系列：真檔一定是 zip，開頭是 b"PK"
_OOXML_EXTS = {".docx", ".xlsx", ".pptx"}


@dataclass
class ValidationResult:
    status: str      # "ok" | "warning" | "failed"
    reason: str      # 中文說明
    suggestion: str = ""  # 修復方向（failed 時才有意義）


def resolve_output_path(raw: str, workflow_name: str = "") -> Path:
    """把 output.path 解析成絕對路徑。規則要跟 runner 的 _resolve_path 一致：

      ~/xxx              → 展開到家目錄
      絕對路徑            → 直接用
      workflows/... 開頭  → 相對資料根目錄
      其他相對路徑        → 相對**這個工作流的**輸出資料夾
    """
    from config import WORKFLOW_DIR

    p = Path(raw).expanduser()
    if p.is_absolute():
        return p
    parts = p.parts
    if parts and parts[0] == "workflows":
        return WORKFLOW_DIR.parent / p
    if workflow_name:
        return WORKFLOW_DIR / workflow_name / p
    return WORKFLOW_DIR / p


def office_format_mismatch(path: Optional[str]) -> Optional[str]:
    """副檔名是 Office（.docx/.xlsx/.pptx）但檔案開頭不是 ZIP magic（PK）→ 回問題描述；
    正常 / 非 Office 副檔名 / 檔案不存在 → None。

    最常見的假檔：把 markdown 或純文字直接存成 .docx（開頭是 '# ' 之類）。
    真正的 OOXML 是 zip，開頭一定是 b'PK'。這個檢查是決定性的。
    """
    if not path:
        return None
    p = Path(path).expanduser()
    if p.suffix.lower() not in _OOXML_EXTS or not p.is_file():
        return None
    try:
        with open(p, "rb") as f:
            head = f.read(4)
    except OSError:
        return None
    if head[:2] != b"PK":
        preview = head.decode("latin-1", "replace").strip()
        return (
            f"輸出檔 {p.name} 副檔名是 Office 格式，但檔案開頭不是 ZIP(PK)、而是「{preview}」"
            f"—— 這不是真正的 {p.suffix.lower()} 檔（很可能是純文字被直接改副檔名）。"
        )
    return None


def deterministic_validate(step, exec_result, logger, workflow_name: str = "") -> ValidationResult:
    """跑完一步之後的確定性檢查。"""
    if exec_result.exit_code != 0:
        return ValidationResult(
            status="failed",
            reason=f"Exit code {exec_result.exit_code}",
            suggestion=(exec_result.stderr or "").strip()[-300:] or "請查看執行 log 取得詳細錯誤",
        )

    if not (step.output and step.output.path):
        return ValidationResult(status="ok", reason="exit code=0")

    p = resolve_output_path(step.output.path, workflow_name)
    if not p.exists():
        return ValidationResult(
            status="failed",
            reason=f"輸出檔案 {step.output.path} 不存在",
            suggestion=f"這步宣告會產出 {step.output.path}，但檔案沒出現。"
                       f"確認腳本真的有寫檔、且路徑與這裡填的一致（解析後：{p}）。",
        )
    if p.is_dir():
        if not any(p.iterdir()):
            return ValidationResult(
                status="failed",
                reason=f"輸出資料夾 {step.output.path} 是空的",
                suggestion="這步宣告的輸出是一個資料夾，但裡面沒有任何檔案。",
            )
        return ValidationResult(status="ok", reason="exit code=0、輸出資料夾非空")

    if p.stat().st_size == 0:
        return ValidationResult(
            status="failed",
            reason=f"輸出檔案 {step.output.path} 為空檔案（0 bytes）",
            suggestion="檔案建立了但沒寫入內容 —— 常見於例外被吞掉、或寫檔的 handle 沒 close。",
        )

    bad_office = office_format_mismatch(str(p))
    if bad_office:
        return ValidationResult(
            status="failed", reason=bad_office,
            suggestion="用 python-docx / openpyxl / python-pptx 產生真正的 Office 檔，"
                       "不要把文字直接存成該副檔名。",
        )

    suffix = p.suffix.lower()
    if suffix == ".csv":
        try:
            with open(p, "r", encoding="utf-8", errors="replace") as f:
                line_count = sum(1 for _ in f)
            if line_count < 2:
                return ValidationResult(
                    status="failed",
                    reason=f"CSV 檔案只有 {line_count} 行（預期至少 header + 1 列資料）",
                    suggestion="確認資料真的有寫進去，不是只寫了標題列。",
                )
        except OSError:
            pass  # 讀不到就不擋 —— 檔案存在且非空已經是通過條件

    elif suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(p, read_only=True)
            sheet_count = len(wb.sheetnames)
            wb.close()
            if sheet_count == 0:
                return ValidationResult(
                    status="failed", reason="Excel 檔案沒有任何工作表",
                    suggestion="",
                )
        except ImportError:
            pass  # 沒裝 openpyxl（它不在必裝清單）→ 跳過這項，不擋
        except Exception:
            pass  # 檔案損毀已被上面的 ZIP magic 檢查涵蓋

    logger.info(f"[{step.name}] ⚡ 確定性檢查通過")
    return ValidationResult(status="ok", reason="exit code=0、輸出檔案存在且非空")
